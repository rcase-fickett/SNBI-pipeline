"""
claude_api.py — Claude API calls with targeted SNBI extraction prompts
"""
import json
import time
import base64
import io
import os
import anthropic

def load_reference_docs(errata_pdf_path=None, crosswalk_path=None):
    """
    Validate reference docs and prepare the crosswalk text.
    Returns dict with keys: errata_pdf_path (str|None), crosswalk_text (str|None).
    """
    docs = {"errata_pdf_path": None, "crosswalk_text": None}

    if errata_pdf_path and os.path.exists(errata_pdf_path):
        docs["errata_pdf_path"] = errata_pdf_path
        print(f"  [ref] Found errata PDF: {os.path.basename(errata_pdf_path)}")
    elif errata_pdf_path:
        print(f"  [ref] Warning: errata PDF not found at {errata_pdf_path}")

    if crosswalk_path and os.path.exists(crosswalk_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(crosswalk_path, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"\n=== {sheet_name} ===")
                for row in ws.iter_rows(values_only=True):
                    if any(c is not None for c in row):
                        parts.append("\t".join("" if c is None else str(c) for c in row))
            docs["crosswalk_text"] = "\n".join(parts)
            print(f"  [ref] Loaded crosswalk: {os.path.basename(crosswalk_path)}")
        except Exception as e:
            print(f"  [ref] Warning: could not load crosswalk: {e}")
    elif crosswalk_path:
        print(f"  [ref] Warning: crosswalk not found at {crosswalk_path}")

    return docs


def render_errata_pages(pdf_path, page_numbers, dpi=150, max_px=1500):
    """
    Render specific 1-indexed pages of the SNBI errata PDF to base64 JPEG images.
    Returns list of (page_num, base64_jpeg_str) tuples.
    """
    import pypdfium2 as pdfium
    from PIL import Image

    images = []
    try:
        doc = pdfium.PdfDocument(pdf_path)
    except Exception as e:
        print(f"  [errata] Could not open PDF: {e}")
        return images

    try:
        total = len(doc)
        for pg_num in page_numbers:
            idx = pg_num - 1
            if idx < 0 or idx >= total:
                continue
            try:
                page = doc[idx]
                scale = dpi / 72.0
                bitmap = page.render(scale=scale, rotation=0)
                pil_img = bitmap.to_pil()
                w, h = pil_img.size
                if max(w, h) > max_px:
                    ratio = max_px / max(w, h)
                    pil_img = pil_img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
                buf = io.BytesIO()
                pil_img.save(buf, format="JPEG", quality=85)
                images.append((pg_num, base64.standard_b64encode(buf.getvalue()).decode()))
            except Exception as e:
                print(f"  [errata] Could not render page {pg_num}: {e}")
    finally:
        doc.close()

    return images


try:
    from lib.snbi_items import (
        PAGE_PLAN, PAGE_SECTION, PAGE_RAIL, PAGE_NOTES,
        PAGE_VICINITY, PAGE_BENT, PAGE_CLEARANCE,
        HIGH, APPROX, FIELD_REQ, NA, PENDING, UNABLE_TO_PARSE,
        ITEMS,
    )
except ImportError:
    from snbi_items import (
        PAGE_PLAN, PAGE_SECTION, PAGE_RAIL, PAGE_NOTES,
        PAGE_VICINITY, PAGE_BENT, PAGE_CLEARANCE,
        HIGH, APPROX, FIELD_REQ, NA, PENDING, UNABLE_TO_PARSE,
        ITEMS,
    )


def _build_item_notes_block(page_type):
    """Build a supplemental guidance block for items relevant to this page type."""
    lines = []
    for item in ITEMS:
        if page_type in item.get("page_types", []) and item.get("notes"):
            lines.append(f"- {item['id']} ({item['name']}): {item['notes']}")
    if not lines:
        return ""
    return "\n\nDetailed SNBI spec guidance for this page type:\n" + "\n".join(lines)


# ── Prompts per page type ──────────────────────────────────────────────────

SYSTEM_PROMPT = """You are an expert bridge engineer extracting SNBI (Specifications for the
National Bridge Inventory) data items from engineering plan drawings.

Return ONLY valid JSON. No preamble, no markdown fences, no explanation outside the JSON.
Use null for value only when the item is truly absent or unreadable.
Round measurements to the nearest tenth of a foot.

For confidence, use:
  HIGH       - value is clearly and unambiguously shown on the drawing
  APPROX     - value was derived, calculated, or estimated from related dimensions
  FIELD_REQ  - a field measurement is required (e.g. drawing is too old or unclear to measure)
  NA         - not applicable per SNBI spec for this bridge type
  UNABLE_TO_PARSE - something is visible (handwriting, stamped note, dimension) but legibility
                    is too poor to read reliably; do NOT use NOT_FOUND in this case

IMPORTANT — always populate "reasoning" even when value is null. Explain what you looked
for, where you looked, and why you could not determine the value. This helps reviewers
decide whether to accept the BrM export value or send the item to field collection.

When SNBI specification pages are included at the start of the message, use them as the
authoritative reference for measurement rules for each item.
"""

PROMPTS = {

PAGE_PLAN: """Extract SNBI data from this bridge Plan & Elevation drawing.
Return JSON with this exact structure:
{
  "bridge_name": null,
  "year_title_block": null,
  "B.G.01": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "B.G.05": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "B.G.13": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "B.G.12": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "B.G.14": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "B.W.01": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "features": [
    {"feature_id": "H01", "location": "C", "name": null},
    {"feature_id": "W01", "location": "B", "name": null}
  ],
  "work_events": [
    {"year": null, "work_codes": [], "reasoning": null}
  ],
  "questions": [],
  "drawing_info": {"title": null, "drawing_number": null, "date": null}
}

Notes:
- B.G.01: Measure between undercopings/faces of end supports along CL. NOT o/o of deck.
- B.G.05: Out-to-out width perpendicular to CL.
- B.G.13: Deck top to water surface or ground (larger value), nearest foot.
- B.W.01: Original construction year from title block. Must be a 4-digit calendar year (e.g. 1987).
  Do NOT report a bridge ID, structure number, drawing number, or any other code — only a year.
  If you cannot find a clear construction year, return null. Flag if this looks like a reconstruction plan.
- work_events: Report ONLY work physically performed on THIS specific bridge structure.
  The source drawing must be a bridge-project-specific sheet (carries this bridge's project
  number or name, is a construction/as-built plan, or has a title block dated AFTER the
  original construction year with a new project description).
  EXCLUDE — do NOT generate work events from:
    • Standard detail sheets or ODOT standard drawings (titles containing "Std.", "Standard",
      "Typical", "General Details", or an agency-wide drawing number)
    • Revision blocks that update a statewide standard (e.g. fence connection hardware changes)
    • Specification sheets, index sheets, or quantity sheets with no structural work callout
  Valid codes (structural/improvement work only):
    BR1=bridge replaced | SP1/SP2/SP3=superstructure replaced/major/minor rehab
    SB1/SB2/SB3=substructure replaced/major/minor rehab | DK1/DK2/DK3=deck replaced/major/minor
    DK4=deck overlaid | JT1/JT2=joints replaced/repaired | BG1/BG2=bearings replaced/repaired
    IP1=widened | IP2=raised | IP3=strengthened | IP4=seismic retrofit
  Do NOT use RT1/RT2 (railing codes) unless a dedicated bridge railing replacement PROJECT
  is clearly documented on a project-specific sheet — fence detail revisions do not qualify.
  When in doubt whether work is bridge-specific vs. a standard detail update, omit it.
- features: List ALL features (carried on AND below). Note if a feature appears to be
  a highway, waterway, or railroad. Flag if widening was proposed but NOT completed.
- questions: List anything ambiguous, conflicting notes, or needing field verification.""",

PAGE_SECTION: """Extract SNBI data from this bridge Typical Section or cross-section drawing.
Return JSON:
{
  "B.G.05": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "B.G.07": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "B.G.08": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "B.G.10": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "B.H.08": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "B.H.16": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "questions": []
}

Notes:
- B.G.07/08: Width from FACE OF RAIL to face of curb. 0.0 if no curb/sidewalk.
  LEFT = south-to-north or west-to-east direction. Check BOTH sides — asymmetric is common.
- B.G.10: 0=no median, 1=open, 2=mountable/flush, 3=non-mountable(barrier/curb >6in).
- B.H.08: Count striped full-width traffic lanes only. Not bike lanes or sidewalks.
- B.H.16: Curb-to-curb width including stabilized shoulders if present.
  Non-mountable curbs >6in high are the limit, not the edge of deck.""",

PAGE_RAIL: """Extract SNBI railing data from this drawing.
Return JSON:
{
  "B.RH.01": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "B.RH.02": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "questions": []
}

Notes:
- Look for NCHRP 350 or MASH test level callouts (e.g. TL-4, TL-3).
- Codes: 3504=NCHRP350 TL-4, 3503=TL-3, M094=MASH TL-4, etc.
- I=unknown/no crash test info. S##=meets agency standard from year ##.
- If rail is a standard ODOT detail drawing with no test level documented, use I.
- Transitions are the connections from approach guardrail to bridge rail.""",

PAGE_NOTES: """Extract SNBI data from these General Notes.
Return JSON:
{
  "B.W.01": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "design_load": {"value": null, "source": null},
  "design_method": {"value": null, "source": null},
  "railroad_service_type": {"value": null, "reasoning": null},
  "navigability_clue": {"value": null, "reasoning": null},
  "work_events": [
    {"year": null, "work_codes": [], "reasoning": null}
  ],
  "questions": []
}

Notes:
- Look for explicit USCG permit references (permit number, "navigable waterway" language, bridge permit stamp). If found, set navigability_clue=Y — this becomes a HIGH override of the Phase 10 APPROX pre-fill. Do NOT guess from waterway size alone.
- Railroad service type: F=freight, P=passenger, M=multiple, I=inactive.
- design_load: HS20, HS25, HL-93, etc.
- design_method: ASD, LFD, LRFD.
- work_events: Same strict rules as the PLAN prompt — only project-specific structural work
  (BR1, SP1-3, SB1-3, DK1-4, JT1-2, BG1-2, IP1-4). Exclude standard detail notes.""",

PAGE_VICINITY: """Extract SNBI data from this Vicinity Map and/or Profile drawing.
Return JSON:
{
  "B.W.01": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "B.G.13": {"value": null, "confidence": null, "reasoning": null, "source": null},
  "feature_name_carried": {"value": null, "source": null},
  "feature_name_crossed": {"value": null, "source": null},
  "work_events": [
    {"year": null, "work_codes": [], "reasoning": null}
  ],
  "questions": []
}

Notes:
- B.G.13: Look for water elevation and deck elevation callouts in the profile.
  Deck EL minus HW EL (or ground EL) = bridge height. Round to nearest foot.
- Feature names: read title block and labels carefully.
- work_events: Only report if the vicinity map or profile clearly shows a construction project
  (e.g. 'widened', 'reconstructed', 'replaced') specific to this bridge. Use codes BR1, SP1-3,
  SB1-3, DK1-4, JT1-2, BG1-2, IP1-4 only. Standard detail revision notes do not qualify.""",

PAGE_BENT: """Extract SNBI substructure data from this bent/abutment drawing.
Return JSON:
{
  "foundation_type": {"value": null, "reasoning": null, "source": null},
  "substructure_material": {"value": null, "reasoning": null},
  "B.N.06_clue": {"value": null, "reasoning": null},
  "questions": []
}

Notes:
- Foundation type: pile bent (timber/steel/concrete), spread footing, drilled shaft, etc.
- B.N.06: Look for fender systems, dolphins, or protective systems around piers in water.""",

PAGE_CLEARANCE: """Extract highway/railroad clearance measurements from this clearance diagram.
These diagrams show vertical and horizontal clearances for features passing BELOW the bridge.
Return JSON:
{
  "features": [
    {
      "feature_name": null,
      "feature_id_guess": null,
      "clearances": {
        "B.H.12": {"value": null, "confidence": null, "reasoning": null},
        "B.H.13": {"value": null, "confidence": null, "reasoning": null},
        "B.H.14": {"value": null, "confidence": null, "reasoning": null},
        "B.H.15": {"value": null, "confidence": null, "reasoning": null},
        "B.RR.02": {"value": null, "confidence": null, "reasoning": null},
        "B.RR.03": {"value": null, "confidence": null, "reasoning": null}
      }
    }
  ],
  "diagram_date": null,
  "questions": []
}

Notes:
- B.H.12: MAXIMUM usable vertical clearance (over 10-ft wide envelope of traveled way).
  This is the BEST clearance available, not the worst.
- B.H.13: MINIMUM vertical clearance anywhere over traveled way + shoulders.
- B.H.14: Left horizontal clearance from edge line to nearest obstruction (0 if 2-way).
- B.H.15: Right horizontal clearance from edge line to nearest obstruction.
- B.RR.02: Top of rail to lowest restriction (for railroad features).
- B.RR.03: CL of track to nearest substructure unit.
- Values highlighted in yellow or boxed are typically the controlling dimensions.
- Convert feet-inches to decimal feet (e.g. 18'-5" = 18.4, always round DOWN per SNBI).""",
}


# ── API client ─────────────────────────────────────────────────────────────

def _call_with_retry(client, **kwargs):
    """Call client.messages.create with exponential backoff on rate limit errors."""
    delays = [30, 60, 120]
    for attempt, wait in enumerate(delays + [None]):
        try:
            return client.messages.create(**kwargs)
        except anthropic.RateLimitError:
            if wait is None:
                raise
            print(f"  [api] Rate limit hit — waiting {wait}s (attempt {attempt + 1}/{len(delays)})...")
            time.sleep(wait)
        except Exception:
            raise


class ClaudeExtractor:
    def __init__(self, api_key, model, delay_sec=1.0):
        self.client = anthropic.Anthropic(api_key=api_key)
        self.model  = model
        self.delay  = delay_sec

    def extract_from_image(self, base64_jpeg, page_type, bridge_id="", geo_context=""):
        """
        Send one page image to Claude and return parsed JSON result.
        Returns (result_dict, raw_text) tuple.
        On error returns ({}, error_message).
        """
        prompt = PROMPTS.get(page_type, PROMPTS[PAGE_PLAN])
        prompt = prompt + _build_item_notes_block(page_type)
        if geo_context:
            prompt = prompt + geo_context

        try:
            response = _call_with_retry(
                self.client,
                model=self.model,
                max_tokens=2000,
                system=SYSTEM_PROMPT,
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type":       "base64",
                                "media_type": "image/jpeg",
                                "data":       base64_jpeg,
                            }
                        },
                        {
                            "type": "text",
                            "text": prompt,
                        }
                    ]
                }]
            )
            raw = response.content[0].text.strip()

            if raw.startswith("```"):
                raw = raw.split("```")[1]
                if raw.startswith("json"):
                    raw = raw[4:]
                raw = raw.strip()

            result = json.loads(raw)
            time.sleep(self.delay)
            return result, raw

        except json.JSONDecodeError as e:
            time.sleep(self.delay)
            return {}, f"JSON parse error: {e}"
        except anthropic.RateLimitError:
            return {}, "Rate limit — giving up after retries"
        except Exception as e:
            time.sleep(self.delay)
            return {}, f"API error: {type(e).__name__}: {e}"


# ── Lesson injection ───────────────────────────────────────────────────────

def build_lesson_block(lessons: dict, page_type: str) -> str:
    """
    Given active lessons dict (item_id → lesson row) and the current page type,
    return a formatted block to prepend to the extraction prompt.
    """
    if not lessons:
        return ""

    PAGE_ITEM_MAP = {
        "PLAN":      ["B.G.01","B.G.05","B.G.13","B.G.12","B.G.14","B.W.01",
                      "B.F.01","B.F.02","B.F.03","B.W.02","B.W.03"],
        "SECTION":   ["B.G.05","B.G.07","B.G.08","B.G.10","B.H.08","B.H.16"],
        "RAIL":      ["B.RH.01","B.RH.02"],
        "NOTES":     ["B.W.01","B.RR.01","B.N.01","B.W.02","B.W.03"],
        "VICINITY":  ["B.W.01","B.G.13","B.F.03","B.W.02","B.W.03"],
        "BENT":      ["B.N.06"],
        "CLEARANCE": ["B.H.12","B.H.13","B.H.14","B.H.15","B.RR.02","B.RR.03"],
    }

    relevant_items = PAGE_ITEM_MAP.get(page_type, [])
    applicable = [
        lessons[iid] for iid in relevant_items
        if iid in lessons and lessons[iid].get("lesson_text")
    ]

    if not applicable:
        return ""

    lines = ["\n\nLESSONS FROM PREVIOUS INSPECTOR REVIEWS (apply these rules):"]
    for lesson in applicable:
        iid = lesson["item_id"]
        text = lesson["lesson_text"]
        n = lesson["correction_count"]
        lines.append(f"\n• {iid} ({n} correction{'s' if n!=1 else ''}):")
        lines.append(f"  {text}")

        try:
            import json as _json
            ex = _json.loads(lesson.get("example_json","{}"))
            if ex and ex.get("value") is not None:
                lines.append(f"  Example correct output: value={ex.get('value')!r}, "
                             f"confidence={ex.get('confidence')!r}")
        except Exception:
            pass

    return "\n".join(lines)


class ClaudeExtractorWithLessons(ClaudeExtractor):
    """
    Extended extractor that injects lessons and targeted SNBI errata pages into prompts.
    Use this in 02_process_bridges.py instead of ClaudeExtractor.
    """
    def __init__(self, api_key, model, delay_sec=1.0, lessons=None,
                 errata_pdf_path=None, crosswalk_text=None):
        super().__init__(api_key, model, delay_sec)
        self.lessons         = lessons or {}
        self.errata_pdf_path = errata_pdf_path
        self.crosswalk_text  = crosswalk_text
        self._errata_cache   = {}  # page_type → list of (pg_num, b64_jpeg)

    def set_lessons(self, lessons: dict):
        self.lessons = lessons

    # Items whose errata pages get sent per page type — limited to items actually
    # extracted in the prompt for that page type, keeping the set small enough for
    # the API (max ~6 pages per call).
    _ERRATA_ITEM_MAP = {
        PAGE_PLAN:      ["B.G.01","B.G.05","B.G.12","B.G.13","B.G.14","B.W.01"],
        PAGE_SECTION:   ["B.G.05","B.G.07","B.G.08","B.G.10","B.H.08","B.H.16"],
        PAGE_RAIL:      ["B.RH.01","B.RH.02"],
        PAGE_NOTES:     ["B.W.01","B.RR.01","B.N.01"],
        PAGE_VICINITY:  ["B.W.01","B.G.13","B.F.03"],
        PAGE_BENT:      ["B.N.06"],
        PAGE_CLEARANCE: ["B.H.12","B.H.13","B.H.14","B.H.15","B.RR.02","B.RR.03"],
    }

    def _get_errata_images(self, page_type):
        """Render and cache SNBI errata pages for items extracted on this page_type."""
        if page_type in self._errata_cache:
            return self._errata_cache[page_type]

        if not self.errata_pdf_path or not os.path.exists(self.errata_pdf_path):
            self._errata_cache[page_type] = []
            return []

        relevant_ids = self._ERRATA_ITEM_MAP.get(page_type, [])
        item_by_id = {i["id"]: i for i in ITEMS}
        page_nums = set()
        for iid in relevant_ids:
            item = item_by_id.get(iid)
            if item:
                page_nums.update(item.get("errata_pages", []))

        if not page_nums:
            self._errata_cache[page_type] = []
            return []

        images = render_errata_pages(self.errata_pdf_path, sorted(page_nums))
        print(f"  [errata] {len(images)} spec page(s) for {page_type}: {sorted(page_nums)}")
        self._errata_cache[page_type] = images
        return images

    def _build_system(self):
        """Build system content list with prompt caching on the crosswalk."""
        blocks = [{"type": "text", "text": SYSTEM_PROMPT}]
        if self.crosswalk_text:
            blocks[-1]["cache_control"] = {"type": "ephemeral"}
            blocks.append({
                "type": "text",
                "text": (
                    "DATA CROSSWALK — BrM column names mapped to SNBI items "
                    "(use this to verify which database field each item comes from):\n"
                    + self.crosswalk_text
                ),
                "cache_control": {"type": "ephemeral"},
            })
        return blocks

    def _build_user_content(self, base64_jpeg, prompt, page_type=None):
        """Build user message content with targeted errata pages before the plan image.

        cache_control is placed only on the last errata image so that all preceding
        errata content is cached as one prefix — the API limit is 4 cache_control
        blocks total (system + user combined), so we use at most 1 here.
        """
        content = []

        if page_type:
            errata_images = self._get_errata_images(page_type)
            if errata_images:
                content.append({
                    "type": "text",
                    "text": "SNBI specification reference pages for the items you will extract:",
                })
                last_idx = len(errata_images) - 1
                for i, (pg_num, img_b64) in enumerate(errata_images):
                    content.append({"type": "text", "text": f"[SNBI Errata — page {pg_num}]"})
                    block = {
                        "type": "image",
                        "source": {"type": "base64", "media_type": "image/jpeg", "data": img_b64},
                    }
                    if i == last_idx:
                        block["cache_control"] = {"type": "ephemeral"}
                    content.append(block)
                content.append({"type": "text", "text": "Bridge plan drawing to analyze:"})

        content.append({
            "type": "image",
            "source": {
                "type":       "base64",
                "media_type": "image/jpeg",
                "data":       base64_jpeg,
            },
        })
        content.append({"type": "text", "text": prompt})
        return content

    def extract_from_image(self, base64_jpeg, page_type, bridge_id="", geo_context=""):
        prompt = PROMPTS.get(page_type, PROMPTS[PAGE_PLAN])
        prompt = prompt + _build_item_notes_block(page_type)

        lesson_block = build_lesson_block(self.lessons, page_type)
        if lesson_block:
            prompt = prompt + lesson_block
        if geo_context:
            prompt = prompt + geo_context

        try:
            response = _call_with_retry(
                self.client,
                model=self.model,
                max_tokens=2000,
                system=self._build_system(),
                messages=[{
                    "role": "user",
                    "content": self._build_user_content(base64_jpeg, prompt, page_type),
                }],
                extra_headers={"anthropic-beta": "prompt-caching-2024-07-31"},
            )
            raw = response.content[0].text.strip()
            if raw.startswith("```"):
                raw = raw.split("```")[1]
                if raw.startswith("json"):
                    raw = raw[4:]
                raw = raw.strip()

            result = json.loads(raw)
            time.sleep(self.delay)
            return result, raw

        except json.JSONDecodeError as e:
            time.sleep(self.delay)
            return {}, f"JSON parse error: {e}"
        except anthropic.RateLimitError:
            return {}, "Rate limit — giving up after retries"
        except Exception as e:
            time.sleep(self.delay)
            return {}, f"API error: {type(e).__name__}: {e}"
